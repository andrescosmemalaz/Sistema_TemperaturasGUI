# Importar la libreria de tkinter
from tkinter import * 
from tkinter import filedialog, messagebox, ttk
import tkinter
#Importar libreria de tkinter para generar calendario en el forms
from tkcalendar import Calendar, DateEntry
#Importar libreria pandastable para el metodo mostrar tabla de informacion desde el programa
from pandastable.core import Table,TableModel
# Importar la libreria openpyxl
import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import workbook
# Importar libreria datetime
from datetime import date, datetime
# Imjportar libreria para comprobar y encontrar la ruta del archivo
import pathlib
#Importar pandas
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class SystemMangementTemperature():
    def __init__(self,root):
        self.root=root
        self.root.title("Sistema Administracion de Temperatura")
        self.root.geometry("350x450+0+0")
        
        title = Label(self.root,text="Sistema Administración Temperatura",bd=12,relief=GROOVE,font=("times new roman",10,"bold"),fg="black")
        title.pack(side=TOP,fill=X)
        
         #Frame de ingreso de Información 
        Managment_Frame = Frame(self.root,bd=4,relief=RIDGE,bg="#D4E1E7")
        Managment_Frame.place(x=20, y=50,width=305,height=380)


        m_title = Label(Managment_Frame,text="Ingresar Temperaturas ", bg="#D4E1E7",fg="black",font=("times new roman",14,"bold"))
        m_title.grid(row=0,columnspan=2,pady=20,padx=20)
        
        #Fecha
        lbl_fecha = Label(Managment_Frame,text="FECHA",bg="#D4E1E7",fg="black",font=("times new roman",9,"bold"))
        lbl_fecha.grid(row=2,column=0, pady=10,padx=20,sticky="w")
        self.fecha = StringVar()
        txt_fecha=DateEntry(Managment_Frame,date_pattern='dd/mm/y',font=("times new roman",5,"bold"),textvariable=self.fecha,state="readonly")
        txt_fecha.grid(row=2, column=1,padx=10,ipadx=20)
        dt =date(2021,12,4)
        txt_fecha.set_date(dt)
        #Temperatura max
        lbl_temp_max = Label(Managment_Frame,text="TEMPERATURA MAX.",bg="#D4E1E7",fg="black",font=("times new roman",7,"bold"))
        lbl_temp_max.grid(row=3,column=0, pady=10,padx=15,sticky="w")
        self.temp_max = StringVar()
        txt_temp_max  =Entry(Managment_Frame,font=("times new roman",7,"bold"),textvariable=self.temp_max)
        txt_temp_max.grid(row=3,column=1,pady=8,padx=14,ipadx=4,sticky="w")
        #Temperatura minima
        lbl_temp_min = Label(Managment_Frame,text="TEMPERATURA MIN.",bg="#D4E1E7",fg="black",font=("times new roman",7,"bold"))
        lbl_temp_min.grid(row=4,column=0, pady=10,padx=15,sticky="w")
        self.temp_min = StringVar()
        txt_temp_min  =Entry(Managment_Frame,font=("times new roman",7,"bold"),textvariable=self.temp_min)
        txt_temp_min.grid(row=4,column=1,pady=8,padx=14,ipadx=4,sticky="w")
        #Temperatura promedio
        lbl_temp_avg = Label(Managment_Frame,text="TEMPERATURA AVG.",bg="#D4E1E7",fg="black",font=("times new roman",7,"bold"))
        lbl_temp_avg.grid(row=5,column=0, pady=10,padx=15,sticky="w")
        self.temp_avg = StringVar()
        txt_temp_avg  =Entry(Managment_Frame,font=("times new roman",7,"bold"),textvariable=self.temp_avg)
        txt_temp_avg.grid(row=5,column=1,pady=8,padx=14,ipadx=4,sticky="w")
        #Tipo de dia
        lbl_vol_tipo_dia = Label(Managment_Frame,text="TIPO DE DIA",bg="#D4E1E7",fg="black",font=("times new roman",7,"bold"))
        lbl_vol_tipo_dia.grid(row=6,column=0, pady=10,padx=15,sticky="w")
        self.tipo_dia = StringVar()
        combo_tipo_dia = ttk.Combobox(Managment_Frame,font=("times new roman",6,"bold"),state="readonly",textvariable=self.tipo_dia)
        combo_tipo_dia['values'] = ('Soleado','Nublado','Lluvioso')
        combo_tipo_dia.grid(row=6,column=1,pady=15,padx=14,sticky="w")
       # Volumen de precipitaciones
        lbl_vol_precipitaciones = Label(Managment_Frame,text="VOL.PRECIPITACIONES",bg="#D4E1E7",fg="black",font=("times new roman",6,"bold"))
        lbl_vol_precipitaciones.grid(row=7,column=0, pady=10,padx=15,sticky="w")
        self.vol_precipitaciones = StringVar()
        txt_vol_precipitaciones  =Entry(Managment_Frame,font=("times new roman",7,"bold"),textvariable=self.vol_precipitaciones)
        txt_vol_precipitaciones.grid(row=7,column=1,pady=8,padx=14,ipadx=4,sticky="w")
        
        #Botones Frame
        btn_frame = Frame(Managment_Frame,bd=3,relief=RIDGE,bg="black")
        btn_frame.place(x=10,y=310,width=240)
        #Botones de agregar, mostrar data en tabla, generar reporte de graficas y limpiar ingreso
        agregar_botom = Button(btn_frame,text="Add",width=6,command=self.agregar).grid(row=0,column=0,padx=5,pady=5)
        mostrar_botom = Button(btn_frame,text="Show data",width=6,command=self.mostrar).grid(row=0,column=1,padx=5,pady=5)
        reporte_botom = Button(btn_frame,text="Reporte",width=6,command=self.reporte).grid(row=0,column=2,padx=5,pady=5)
        limpiar_botom = Button(btn_frame,text="Clean",width=6,command=self.limpiar_texto).grid(row=0,column=3,padx=5,pady=5)
        
        
    #EDefiniendo archivo excel y metodo active para usar excel
    file = pathlib.Path("F:/andres/Ciclo 10/Topicos/Python Clases/Tkinter/PC-4/Temperaturadatos.xlsx")
    wb = load_workbook(file)  
    sheet = wb.active
    df = pd.read_excel(file)
    df2 = pd.read_excel(file)
    fecha_list = df2['FECHA'].tolist()
    for x in fecha_list:
        fecha = str(x)
        fecha_list.append(fecha)
    
    def archivo_excel(self):
        self.sheet.column_dimensions['A'].width = 30
        self.sheet.column_dimensions['B'].width = 30
        self.sheet.column_dimensions['C'].width = 30
        self.sheet.column_dimensions['D'].width = 30
        self.sheet.column_dimensions['E'].width = 30
        self.sheet.column_dimensions['F'].width = 30

        self.sheet.cell(row=1, column=1).value = "FECHA"
        self.sheet.cell(row=1, column=2).value = "TEMPERATURA_MAXIMA"
        self.sheet.cell(row=1, column=3).value = "TEMPERATURA_MINIMA"
        self.sheet.cell(row=1, column=4).value = "TEMPERATURA_PROMEDIO"
        self.sheet.cell(row=1, column=5).value = "TIPO_DIA"
        self.sheet.cell(row=1, column=6).value = "VOLUMEN_PRECIPITACIONES"
    #Limpiar ingreso en pantalla
    def limpiar_texto(self):
        self.fecha.set("")
        self.temp_max.set("")
        self.temp_min.set("")
        self.temp_avg.set("")
        self.tipo_dia.set("")
        self.vol_precipitaciones.set("")
    #Metodo para agregar data en cada fila
    def agregar(self):
        x = self.fecha.get()
        if x in self.fecha_list:
            tkinter.messagebox.showwarning(title="Alerta", message="La fecha ya esta registrada",detail="Ingresé otro fecha")
            self.fecha.set("")
            self.temp_max.set("")
            self.temp_min.set("")
            self.temp_avg.set("")
            self.tipo_dia.set("")
            self.vol_precipitaciones.set("")
        else:
            temp_text1 = self.fecha.get()
            temp_text2 = self.temp_max.get()
            temp_text3 = self.temp_min.get()
            temp_text4 = self.temp_avg.get()
            temp_text5 = self.tipo_dia.get()
            temp_text6 = self.vol_precipitaciones.get()
        
            current_row = self.sheet.max_row
            current_column = self.sheet.max_column 
            self.sheet.cell(row=current_row + 1, column=1).value = self.fecha.get()
            self.sheet.cell(row=current_row + 1, column=2).value = self.temp_max.get()
            self.sheet.cell(row=current_row + 1, column=3).value = self.temp_min.get()
            self.sheet.cell(row=current_row + 1, column=4).value = self.temp_avg.get()
            self.sheet.cell(row=current_row + 1, column=5).value = self.tipo_dia.get()
            self.sheet.cell(row=current_row + 1, column=6).value = self.vol_precipitaciones.get()
            self.wb.save(self.file)  
            self.limpiar_texto()
    #Metodo para mostrar la informacion del excel en una tabla interactiva
    def mostrar(self):
        df = pd.read_excel(self.file)
        df['FECHA'] = pd.to_datetime(df['FECHA'])
        df['TEMPERATURA_MAXIMA'] = pd.to_numeric(df['TEMPERATURA_MAXIMA'])
        df['TEMPERATURA_MINIMA'] = pd.to_numeric(df['TEMPERATURA_MINIMA'])
        df['TEMPERATURA_PROMEDIO'] = pd.to_numeric(df['TEMPERATURA_PROMEDIO'])
        df['VOLUMEN_PRECIPITACIONES'] = pd.to_numeric(df['VOLUMEN_PRECIPITACIONES'])
        self.root = Tk()
        self.root.title("Tabla de Datos")
        title = Label(self.root,text="Data del Sistema de Temperatura",bd=12,relief=GROOVE,font=("times new roman",10,"bold"),fg="black")
        title.pack(side=TOP,fill=X)
        self.root.geometry("720x450+0+0")
        self.Extra_Frame = Frame(self.root,bd=4,relief=RIDGE,bg="#D4E1E7")
        self.Extra_Frame.place(x=40,y=50,width=650,height=380)
        self.pt = Table(self.Extra_Frame,dataframe=df)
        self.pt.show()
    #Generar  Graficas 
    def reporte(self):
        df = pd.read_excel(self.file)
        df['FECHA'] = pd.to_datetime(df['FECHA'])
        df['TEMPERATURA_MAXIMA'] = pd.to_numeric(df['TEMPERATURA_MAXIMA'])
        df['TEMPERATURA_MINIMA'] = pd.to_numeric(df['TEMPERATURA_MINIMA'])
        df['TEMPERATURA_PROMEDIO'] = pd.to_numeric(df['TEMPERATURA_PROMEDIO'])
        df['VOLUMEN_PRECIPITACIONES'] = pd.to_numeric(df['VOLUMEN_PRECIPITACIONES'])
        self.root = Tk()
        self.root.title("Tabla de Datos")
        self.root.geometry("990x325")
        self.Graficas_Frame1 = Frame(self.root,bd=4,relief=RIDGE,bg="#D4E1E7")
        self.Graficas_Frame1.pack()
        fecha = df['FECHA']
        temperatura_max = df['TEMPERATURA_MAXIMA']
        temperatura_min = df['TEMPERATURA_MINIMA']
        temperatura_avg = df['TEMPERATURA_PROMEDIO']
        volumen_precipitaciones = df['VOLUMEN_PRECIPITACIONES']
        fig1, axs1 = plt.subplots(3,1 , dpi=80 ,figsize=(13, 4), 
            sharey=True, facecolor='#00f9f844')

        fig1.suptitle('Graficas Matplotlib')

        
        axs1[0].plot(fecha,temperatura_max, color = 'y')
        axs1[1].plot(fecha, temperatura_min, color = 'g')
        axs1[2].plot(fecha, temperatura_avg, color = 'm')
        canvas1 = FigureCanvasTkAgg(fig1, master =  self.Graficas_Frame1)  # Crea el area de dibujo en Tkinter
        canvas1.draw()
        canvas1.get_tk_widget().grid(column=0, row=0)
        

    

if __name__ == "__main__":
        root = Tk()
        obj = SystemMangementTemperature(root)
        root.resizable(width=False, height=False)
        root.mainloop()
 