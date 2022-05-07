from tkinter import *
from tkinter import font
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pathlib

main = Tk()
main.title("Andres.cosme - OpenPyXL")
main.geometry("250x400")

file = pathlib.Path("Temperatura.xlsx")
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Username"
    sheet['C1'] = "Password"
    sheet['D1'] = "Mail"
    sheet['E1'] = "Gender"
    sheet['F1'] = "Subscrption"
    
    file.save("Temperatura.xlsx")


def show_Password():
    if var.get()==1:
        passEntry.config(show='')
    if var.get()==0:
        passEntry.config(show='')


def submit():
    y = name.get()
    z = user.get()
    z1 = passEntry.get()
    y1 = emailEntry.get()
    print(y)
    print(z)
    print(z1)
    print(y1)

    file = openpyxl.load_workbook("Temperatura.xlsx")
    sheet = file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=y)
    sheet.cell(column=2,row=sheet.max_row,value = z)
    sheet.cell(column=3,row=sheet.max_row,value=z1)
    sheet.cell(column=4,row=sheet.max_row,value=y1)
    
    
    if var1.get() == 1:
        gen = "Male"
        print("Male")
        sheet.cell(column=5,row=sheet.max_row,value="Male")
    else:
        gen = "Female"
        print("Female")
        sheet.cell(column=5,row=sheet.max_row,value="Female")
    
    if var2.get() == 1:
        print("Standard")
        sheet.cell(column=6,row=sheet.max_row,value="Standard")
    if var3.get() == 1 :
        print("Premium")
        sheet.cell(column=6,row=sheet.max_row,value="Premium")
    file.save("Temperatura.xlsx")


frame1 = LabelFrame(main,text="Login Detalles").pack(expand='yes',fill='both')

Label(frame1,text="Nombre:").place(x=50,y=30)
Label(frame1,text="Nombreuser").place(x=50,y=70)
Label(frame1,text="Password:").place(x=50,y=110)   
Label(frame1,text="Mail ID:").place(x=50,y=150)

frame2 = LabelFrame(main, text='Otros Detalles').pack(expand='yes', fill='both')

name = Entry(frame1)
name.place(x=250,y=30)
user = Entry(frame1)
user.place(x=250,y=70)
password = StringVar()
passEntry = Entry(frame1,textvariable=password, show='*')
passEntry.place(x=250,y=110)

emailEntry = Entry(frame1)
emailEntry.place(x=250,y=150,width=250)

var = IntVar()
var.set('0')
chkbutton = Checkbutton(frame1,text='Mostrar',variable=var, onvalue=1,offvalue=0,command=show_Password).place(x=350,y=110)

label_3 = Label(frame2, text='Genero',width=20,font=('bold',10))
label_3.place(x=50,y=300)
var1 = IntVar()
Radiobutton(frame2,text="Masculino",padx=5,variable=var1,value=1).place(x=200,y=300)
Radiobutton(frame2,text="Femenino",padx=20,variable=var1,value=2).place(x=280,y=300)

label_4 = Label(frame2,text='Suscripción',width=20,font=("bold",10))
label_4.place(x=50,y=360)

var2= IntVar()
Checkbutton(frame2,text="Estandar",variable=var2).place(x=200,y=360)
var3=IntVar()
Checkbutton(frame2,text="Premium",command=var3).place(x=300,y=360)

Button(frame2,text="Suscrbirse",command=submit).place(x=400,y=400)



main.mainloop()